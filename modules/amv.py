"""活跃市值（AMV）多空区间：选股的总开关。

活跃市值是本系统判断"今天能不能选股"的**唯一**依据。全市场宽度
（``market_context``）不再承担这个角色，降级为辅助建仓参考。

区间规则
--------

设 ``p`` 为当日涨幅（百分数），``p_prev`` 为前一交易日涨幅：

    若 p < -2.3                          → 空头区间
    否则若 p >= 4 或 (p + p_prev) >= 4   → 多头区间
    否则                                  → 沿用前一日区间

两条容易写错的细节，都是拿 8180 个交易日的官方标注反推出来的：

1. **空头触发优先于两日累计多头。** 1993-03-08 涨 10.77% 进多头，次日跌 5.44%，
   两日累计仍有 +5.33%，但标注是空头——单日暴跌直接压过累计涨幅。
   先判空头、再判多头，顺序反了会有 18 天判错。
2. **涨幅必须用收盘价现算，不能用四舍五入到两位小数的显示值。**
   1993-10-11 显示 -2.30%（实为 -2.295082%，多头）与 1997-11-11 显示 -2.30%
   （实为 -2.302724%，空头）差在万分之一上。用显示值会有 10 天判错。

按上述规则回放 1993-01-04 ~ 2026-08-07 共 8180 个交易日，与官方标注
**逐日 100% 吻合**（见 tests/test_amv.py 的回归测试）。

数据来源
--------

- 每日（当前主路径）：``scripts/sync_amv.py`` 从百度网盘分享链接拉当天的表
  再整表导入，由 systemd timer ``quant-monitor-amv.timer`` 触发。
- 历史/手工：``zt amv import <文件>``，csv / xlsx / zip 都认。整表 upsert，
  重复导入同一份文件是幂等的。
- 兜底：``zt amv add <日期> --close <收盘价>``。也接受 ``--pct``，但**精度不足**
  ——落在 -2.3% 边界附近时结论可能相反，所以能给收盘价就给收盘价。
"""

from __future__ import annotations

import csv
import logging
import re
import shutil
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from .database import get_connection

logger = logging.getLogger(__name__)


# ==================== 区间常量 ====================
#
# 这三个数字是用户给定的交易规则，不是我调出来的参数。
# 拿 8180 个交易日的官方标注验证过，改动任何一个都会偏离标注。

# 进入多头：单日涨幅 >= 4%，或连续两日累计涨幅 >= 4%
BULL_THRESHOLD = 4.0
# 累计窗口的天数（"连续 1 天或两天"）
BULL_WINDOW = 2
# 进入空头：单日跌幅**超过** 2.3%（严格小于，-2.3% 整不触发）
BEAR_THRESHOLD = -2.3

REGIME_BULL = "多头区间"
REGIME_BEAR = "空头区间"


@dataclass
class AmvDay:
    trade_date: str
    close: float
    pct_chg: float | None
    regime: str
    regime_imported: str = ""

    @property
    def can_select(self) -> bool:
        """该区间下是否允许选股/新建仓。"""
        return self.regime == REGIME_BULL


def _norm_date(value: str) -> str:
    """把 1993-01-04 / 19930104 / 1993/1/4 统一成 YYYYMMDD。"""
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) != 8:
        raise ValueError(f"无法解析日期: {value!r}")
    return digits


def _parse_pct(value: str | float | None) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().rstrip("%")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


# ==================== 区间状态机 ====================


def classify(pcts: Iterable[float | None], initial: str | None = None) -> list[str]:
    """按日涨幅序列回放多空区间。

    Args:
        pcts: 按时间升序的日涨幅（百分数）；首日可为 None
        initial: 序列开始前的区间状态，None 表示未定

    Returns:
        与输入等长的区间列表；状态尚未确定时为空字符串
    """
    out: list[str] = []
    state = initial
    prev: float | None = None
    for p in pcts:
        if p is not None:
            # 顺序不能反：单日暴跌压过两日累计涨幅
            if p < BEAR_THRESHOLD:
                state = REGIME_BEAR
            elif p >= BULL_THRESHOLD or (prev is not None and p + prev >= BULL_THRESHOLD):
                state = REGIME_BULL
            prev = p
        out.append(state or "")
    return out


def recompute_regimes() -> int:
    """用收盘价重算全序列的涨幅与区间，写回库。

    涨幅一律由收盘价现算而不信任导入值：导入的 CSV 只给到两位小数，
    在 -2.3% 边界上会翻车（详见模块文档）。

    Returns:
        更新的行数
    """
    with get_connection() as conn:
        rows = conn.execute("SELECT trade_date, close FROM amv_daily ORDER BY trade_date").fetchall()
        if not rows:
            return 0

        dates = [str(r[0]) for r in rows]
        closes = [float(r[1]) for r in rows]
        pcts: list[float | None] = [None]
        for i in range(1, len(closes)):
            prev = closes[i - 1]
            pcts.append((closes[i] - prev) / prev * 100 if prev else None)

        regimes = classify(pcts)
        conn.executemany(
            "UPDATE amv_daily SET pct_chg = ?, regime = ? WHERE trade_date = ?",
            [(p, r, d) for d, p, r in zip(dates, pcts, regimes)],
        )
    return len(dates)


# ==================== 表格解析 ====================
#
# 活跃市值的原始表由用户从行情终端导出后放百度网盘，格式不受我们控制：
# 见过 CSV，也可能是 xlsx，或者一个装着若干 CSV 的 zip（.gitignore 里
# 那条「【批量下载】*.zip」就是这么来的）。与其每换一次导出方式改一次代码，
# 不如在这里一次把三种容器和常见列名都认下来。

# 列名别名。同一列在不同导出器里叫法不同，全部认下来。
_DATE_KEYS = ("date", "trade_date", "日期", "交易日期", "时间")
_OPEN_KEYS = ("open", "开盘", "开盘价")
_HIGH_KEYS = ("high", "最高", "最高价")
_LOW_KEYS = ("low", "最低", "最低价")
_CLOSE_KEYS = ("close", "收盘", "收盘价", "最新", "最新价")
_VOLUME_KEYS = ("volume", "vol", "成交量")
_AMOUNT_KEYS = ("amount", "成交额", "成交金额")
_REGIME_KEYS = ("区间", "regime", "多空区间")

# zip 里认哪些后缀当表格。其余（说明文件、图片）跳过。
_TABLE_SUFFIXES = (".csv", ".txt", ".xlsx", ".xlsm")


def _norm_key(key: Any) -> str:
    """表头去空白、去 BOM。"""
    return str(key if key is not None else "").strip().lstrip("﻿")


def _pick(row: dict[str, Any], keys: tuple[str, ...]) -> Any:
    """按别名顺序取第一个非空值。"""
    for k in keys:
        v = row.get(k)
        if v is not None and str(v).strip() != "":
            return v
    return None


def _to_float(value: Any) -> float | None:
    """单元格转 float；空值、非数字一律 None（不抛异常）。"""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").rstrip("%")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _cell_to_date(value: Any) -> str:
    """单元格转 YYYYMMDD。

    Excel 的日期列 openpyxl 会还原成 datetime 对象，``str()`` 出来是
    ``2026-08-25 00:00:00``，交给 _norm_date 会数出 14 位数字然后报
    「无法解析日期」——错误信息指向格式非法，实际是类型没转。
    """
    if isinstance(value, (datetime, date)):  # datetime 是 date 的子类，一起接住
        return value.strftime("%Y%m%d")
    return _norm_date(value)


def _decode(raw: bytes) -> str:
    """CSV 解码：先 UTF-8 再 GBK。

    行情终端导出的 CSV 有相当比例是 GBK。原来固定 ``utf-8-sig`` +
    ``errors="replace"`` 解，GBK 文件**不会报错**，而是整行表头变成乱码，
    于是一列都匹配不上，最后抛「没有解析出任何有效行情行」——
    错误信息指向文件是空的，实际是编码猜错了。
    """
    for enc in ("utf-8-sig", "gbk"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8-sig", errors="replace")


def _read_csv(path: Path) -> list[dict[str, Any]]:
    return [
        {_norm_key(k): v for k, v in row.items()} for row in csv.DictReader(_decode(path.read_bytes()).splitlines())
    ]


def _read_excel(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - 依赖缺失时才走到
        raise RuntimeError(f"读 {path.suffix} 需要 openpyxl：pip install openpyxl") from exc

    # data_only=True 取公式的缓存值；否则带公式的单元格读出来是 "=A1*2" 这种字符串。
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return []
        keys = [_norm_key(h) for h in header]
        out = []
        for values in rows:
            if all(v is None or str(v).strip() == "" for v in values):
                continue  # 表尾的空行
            out.append(dict(zip(keys, values)))
        return out
    finally:
        wb.close()


def _read_zip(path: Path) -> list[dict[str, Any]]:
    """解开压缩包，把里面所有表格按文件名排序后依次读出来。

    Windows/网盘打的包，成员文件名多半是 GBK 且没置 UTF-8 标志位，
    zipfile 会按 cp437 解成乱码。乱码名不影响解压内容，但会让
    「挑出 .csv」这步失灵——包里明明有表格，却报「没有表格文件」。
    """
    rows: list[dict[str, Any]] = []
    with zipfile.ZipFile(path) as zf:
        members = []
        for info in zf.infolist():
            if info.is_dir():
                continue
            name = info.filename
            if not info.flag_bits & 0x800:  # 0x800 = 文件名是 UTF-8
                try:
                    name = name.encode("cp437").decode("gbk")
                except (UnicodeEncodeError, UnicodeDecodeError):
                    pass
            if Path(name).suffix.lower() in _TABLE_SUFFIXES:
                members.append((name, info))

        if not members:
            raise ValueError(f"{path} 里没有 {'/'.join(_TABLE_SUFFIXES)} 表格文件")

        with tempfile.TemporaryDirectory() as tmp:
            for name, info in sorted(members):
                dest = Path(tmp) / Path(name).name
                with zf.open(info) as src, dest.open("wb") as out:
                    shutil.copyfileobj(src, out)
                rows.extend(_read_table(dest))
    return rows


def _read_table(path: Path) -> list[dict[str, Any]]:
    """按后缀分派到具体读法。zip 会递归回到这里。"""
    suffix = path.suffix.lower()
    if suffix == ".zip":
        return _read_zip(path)
    if suffix in (".xlsx", ".xlsm"):
        return _read_excel(path)
    if suffix == ".xls":
        # openpyxl 只认 OOXML，老式 BIFF 的 .xls 读出来是 InvalidFileException，
        # 报错文本是英文的 "openpyxl does not support the old .xls format"，
        # 不给出路会让人以为是文件坏了。
        raise ValueError(
            f"{path} 是老式 .xls（BIFF）格式，openpyxl 读不了。先转一道：soffice --headless --convert-to xlsx '{path}'"
        )
    return _read_csv(path)


# ==================== 导入与录入 ====================


def import_history(source: str | Path, *, dry_run: bool = False) -> dict[str, Any]:
    """导入活跃市值表格（csv / xlsx / zip 都认）。

    保留原始的「区间」列到 regime_imported，供回归测试比对；
    实际生效的 regime 由 recompute_regimes 按规则重算。

    整表 upsert，所以重复导入同一份文件是幂等的；每日下载的全量表
    直接喂进来即可，不需要先切出增量。

    Args:
        source: 文件路径，后缀决定读法
        dry_run: 只解析不落库，用来在换了导出格式后先核对一遍列名映射

    Returns:
        imported/skipped/start/end/columns/preview，dry_run 时 imported 是「将要写入」的行数
    """
    path = Path(source)
    if not path.exists():
        raise FileNotFoundError(f"找不到文件: {path}")

    raw_rows = _read_table(path)
    columns = sorted({k for row in raw_rows for k in row if k})

    records = []
    skipped = 0
    for row in raw_rows:
        raw_date = _pick(row, _DATE_KEYS)
        close = _to_float(_pick(row, _CLOSE_KEYS))
        if raw_date is None or close is None or close <= 0:
            # 补录用的占位行（只有日期没有收盘价）走这里，属于正常跳过。
            skipped += 1
            continue
        try:
            trade_date = _cell_to_date(raw_date)
        except ValueError:
            skipped += 1
            continue
        records.append(
            (
                trade_date,
                _to_float(_pick(row, _OPEN_KEYS)),
                _to_float(_pick(row, _HIGH_KEYS)),
                _to_float(_pick(row, _LOW_KEYS)),
                close,
                _to_float(_pick(row, _VOLUME_KEYS)),
                _to_float(_pick(row, _AMOUNT_KEYS)),
                str(_pick(row, _REGIME_KEYS) or "").strip(),
            )
        )

    if not records:
        # 一行都没解析出来，最常见的两个原因是列名对不上和编码猜错，
        # 光说"没有有效行"会让人去查文件是不是空的，所以把表头一并打出来。
        raise ValueError(
            f"{path} 中没有解析出任何有效行情行（读到 {len(raw_rows)} 行，"
            f"表头 {columns or '空'}）。日期列需叫 {'/'.join(_DATE_KEYS)} 之一，"
            f"收盘列需叫 {'/'.join(_CLOSE_KEYS)} 之一"
        )

    # zip 里多份文件拼起来时顺序不一定按日期，start/end 和 upsert 都依赖有序。
    # sort 是稳定的，同一天出现多次时仍保留文件先后，后写的覆盖先写的。
    records.sort(key=lambda r: r[0])

    result = {
        "imported": len(records),
        "skipped": skipped,
        "start": records[0][0],
        "end": records[-1][0],
        "source": str(path),
        "columns": columns,
        "dry_run": dry_run,
        "preview": [{"trade_date": r[0], "close": r[4], "regime_imported": r[7]} for r in (records[:3] + records[-3:])],
    }
    if dry_run:
        return result

    with get_connection() as conn:
        conn.executemany(
            """
            INSERT INTO amv_daily
              (trade_date, open, high, low, close, volume, amount, regime_imported)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(trade_date) DO UPDATE SET
                open = COALESCE(excluded.open, amv_daily.open),
                high = COALESCE(excluded.high, amv_daily.high),
                low = COALESCE(excluded.low, amv_daily.low),
                close = excluded.close,
                volume = COALESCE(excluded.volume, amv_daily.volume),
                amount = COALESCE(excluded.amount, amv_daily.amount),
                -- 空值不覆盖旧值。每天下的那份表只有 OHLCV，没有「区间」列，
                -- 直接 excluded.regime_imported 会把历史「增强」表导进来的
                -- 8180 行官方标注一次性清空——那是校验区间规则的唯一地面真值，
                -- 清掉之后 zt amv verify 永远返回 0/0，而且没有任何报错。
                regime_imported = COALESCE(NULLIF(excluded.regime_imported, ''), amv_daily.regime_imported)
            """,
            records,
        )

    recompute_regimes()
    return result


def add_daily(
    trade_date: str,
    *,
    close: float | None = None,
    pct_chg: float | None = None,
) -> AmvDay:
    """录入单日活跃市值（收盘后由用户提供）。

    close 与 pct_chg 至少给一个。**优先给 close**——只给涨幅时，
    若它是四舍五入过的百分数，在 -2.3% 边界上可能得出相反的区间结论。
    只给 pct_chg 时，用它和上一日收盘反推一个收盘价存进去，
    保证 recompute_regimes 的口径统一。
    """
    date = _norm_date(trade_date)
    if close is None and pct_chg is None:
        raise ValueError("close 与 pct_chg 至少提供一个")

    with get_connection() as conn:
        row = conn.execute(
            "SELECT close FROM amv_daily WHERE trade_date < ? ORDER BY trade_date DESC LIMIT 1", (date,)
        ).fetchone()
        prev_close = float(row[0]) if row else None

        if close is None:
            if prev_close is None:
                raise ValueError(f"{date} 之前没有任何活跃市值数据，只给涨幅无法反推收盘价，请提供 --close")
            close = prev_close * (1 + pct_chg / 100.0)
            logger.warning(
                "只提供了涨幅 %.4f%%，已按上一日收盘 %.4f 反推收盘价 %.4f；"
                "涨幅若是四舍五入值，在 -2.3%% 边界附近可能影响区间判定",
                pct_chg,
                prev_close,
                close,
            )

        conn.execute(
            """
            INSERT INTO amv_daily (trade_date, close)
            VALUES (?, ?)
            ON CONFLICT(trade_date) DO UPDATE SET close = excluded.close
            """,
            (date, float(close)),
        )

    recompute_regimes()
    day = get_day(date)
    if day is None:  # 理论上不可能：刚写进去
        raise RuntimeError(f"{date} 写入后读取失败")
    return day


# ==================== 查询 ====================


def get_day(trade_date: str) -> AmvDay | None:
    """精确取某一天。"""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT trade_date, close, pct_chg, COALESCE(regime,''), COALESCE(regime_imported,'') "
            "FROM amv_daily WHERE trade_date = ?",
            (_norm_date(trade_date),),
        ).fetchone()
    if row is None:
        return None
    return AmvDay(str(row[0]), float(row[1]), row[2], str(row[3]), str(row[4]))


def get_regime(trade_date: str | None = None) -> AmvDay | None:
    """取截至 trade_date 的最近一条活跃市值记录（含区间）。

    不要求精确匹配：活跃市值由用户手工录入，可能比行情库晚一天。
    向前回退到最近一条，调用方可通过 trade_date 字段看出用的是哪天的数据。
    """
    with get_connection() as conn:
        if trade_date:
            row = conn.execute(
                "SELECT trade_date, close, pct_chg, COALESCE(regime,''), COALESCE(regime_imported,'') "
                "FROM amv_daily WHERE trade_date <= ? ORDER BY trade_date DESC LIMIT 1",
                (_norm_date(trade_date),),
            ).fetchone()
        else:
            row = conn.execute(
                "SELECT trade_date, close, pct_chg, COALESCE(regime,''), COALESCE(regime_imported,'') "
                "FROM amv_daily ORDER BY trade_date DESC LIMIT 1"
            ).fetchone()
    if row is None:
        return None
    return AmvDay(str(row[0]), float(row[1]), row[2], str(row[3]), str(row[4]))


def recent(limit: int = 20, end_date: str | None = None) -> list[AmvDay]:
    sql = (
        "SELECT trade_date, close, pct_chg, COALESCE(regime,''), COALESCE(regime_imported,'') FROM amv_daily "
        + ("WHERE trade_date <= ? " if end_date else "")
        + "ORDER BY trade_date DESC LIMIT ?"
    )
    params = (_norm_date(end_date), limit) if end_date else (limit,)
    with get_connection() as conn:
        rows = conn.execute(sql, params).fetchall()
    return [AmvDay(str(r[0]), float(r[1]), r[2], str(r[3]), str(r[4])) for r in reversed(rows)]


def regime_segments(limit: int = 20) -> list[dict[str, Any]]:
    """最近 limit 段多空区间（起止日期与天数）。"""
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT trade_date, COALESCE(regime,'') FROM amv_daily WHERE regime != '' ORDER BY trade_date"
        ).fetchall()
    segments: list[dict[str, Any]] = []
    # 循环变量不能再叫 date：模块顶部导入了 datetime.date（xlsx 的日期单元格要用），
    # 叫 date 会把它遮掉（ruff F402）。
    for trade_date, regime in rows:
        if segments and segments[-1]["regime"] == regime:
            segments[-1]["end"] = str(trade_date)
            segments[-1]["days"] += 1
        else:
            segments.append({"regime": str(regime), "start": str(trade_date), "end": str(trade_date), "days": 1})
    return segments[-limit:]


def verify_against_imported() -> dict[str, Any]:
    """把重算的区间与导入的官方标注逐日比对（回归用）。"""
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT trade_date, COALESCE(regime,''), COALESCE(regime_imported,'') "
            "FROM amv_daily WHERE regime_imported != '' ORDER BY trade_date"
        ).fetchall()
    mismatches = [
        {"trade_date": str(d), "computed": str(c), "imported": str(i)} for d, c, i in rows if str(c) != str(i)
    ]
    total = len(rows)
    return {
        "total": total,
        "matched": total - len(mismatches),
        "mismatches": mismatches,
        "accuracy": round((total - len(mismatches)) / total * 100, 4) if total else 0.0,
    }


def format_amv_status(day: AmvDay | None, segments: list[dict[str, Any]] | None = None) -> str:
    """当前区间状态的人类可读输出。"""
    if day is None:
        return "活跃市值库为空。先 `zt amv import <csv>` 导入历史，再用 `zt amv add` 逐日录入。"

    icon = "可选股" if day.can_select else "停止选股"
    lines = [
        "=" * 62,
        f"活跃市值 {day.trade_date}   收盘 {day.close:,.2f}"
        + (f"   涨幅 {day.pct_chg:+.4f}%" if day.pct_chg is not None else ""),
        f"区间: {day.regime or '未定'}   →   {icon}",
        "=" * 62,
        f"规则: 单日跌幅 < {BEAR_THRESHOLD}% → 空头；单日或连续两日累计涨幅 ≥ {BULL_THRESHOLD}% → 多头；否则沿用",
    ]
    if segments:
        lines.append("\n【最近区间】")
        for s in segments:
            lines.append(f"  {s['regime']}  {s['start']} ~ {s['end']}  ({s['days']} 个交易日)")
    return "\n".join(lines)
